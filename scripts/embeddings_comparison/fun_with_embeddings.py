from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sentence_transformers import SentenceTransformer


# tnx ChatGPT for this code

def get_current_datetime():
    formatted_datetime = datetime.now().strftime("%Y_%m_%d__%H_%M_%S")
    return formatted_datetime


def calculate_dot_product(questions, answers, sentence_transformer, path_to_file):
    # Calculate embeddings for questions and answers
    question_embeddings = sentence_transformer.encode(questions, convert_to_tensor=True)
    answer_embeddings = sentence_transformer.encode(answers, convert_to_tensor=True)

    # Normalize embeddings
    normalized_question_embeddings = question_embeddings / np.linalg.norm(question_embeddings, axis=1)[:, np.newaxis]
    normalized_answer_embeddings = answer_embeddings / np.linalg.norm(answer_embeddings, axis=1)[:, np.newaxis]

    # Calculate dot product between question and answer embeddings
    dot_product_matrix = np.dot(normalized_question_embeddings, normalized_answer_embeddings.T)

    # Create a new Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active

    # Write the labels (questions and answers) to the first row and column
    for i, question in enumerate(questions):
        sheet.cell(row=i + 2, column=1).value = question
    for j, answer in enumerate(answers):
        sheet.cell(row=1, column=j + 2).value = answer

    # Define fill colors for different ranges of dot product scores
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Write the dot product scores to the corresponding cells
    for i in range(len(questions)):
        for j in range(len(answers)):
            cell = sheet.cell(row=i + 2, column=j + 2)
            dot_product_score = dot_product_matrix[i, j]
            cell.value = dot_product_score
            if dot_product_score > 0.6:
                cell.fill = green_fill
            elif dot_product_score >= 0.2:
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill

    # Format the width of the cells
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Save the workbook to an Excel file
    workbook.save(path_to_file)


MODEL = 'paraphrase-multilingual-mpnet-base-v2'
# MODEL = 'all-MiniLM-L6-v2'
embedder = SentenceTransformer(MODEL)

# Сделать матрицу кто должен на что триггериться
# Сделать скор - процент попаданий (модель + трешхолд)

example_questions = [
    'When is my dentist appointment?',
    'When did I last defrosted my fridge?',
    'Когда я последний раз стирал белье постельное',
    'что мне осталось купить?',
    'на какой серии игры престолов я остановился?',
    'How much did I spend in a month?',
    'Что я напланировал на неделю?',
    'When is Bob\'s birthday?',
]

example_answers = [
    'I have a dentist appointment tomorrow at 10 AM.',
    'У меня завтра запись к стоматологку завтра в 10 утра',
    'Разморозил сегодня холодильник',
    'I defrosted the fridge today',
    'Постирал белье',
    'I need to buy milk, eggs and sugar',
    'нужно купить молока, яиц и сахара',
    'игра престолов остановился на 4 серии 24:32',
    'купил носки за 699р',
    'Bob birthday 10 april',
]

file = fr'.\reports\{MODEL}__{get_current_datetime()}.xlsx'
# Calculate dot product scores and save them in an Excel file
calculate_dot_product(example_questions, example_answers, embedder, file)
